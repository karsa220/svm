import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox
import traceback


# ================= 各个模块的处理函数 =================

def process_rw01_02(df, wb):
    num_end = int(df['年末的人'].sum())
    num_avg = round(df['总月份数（正式职工年平均人数=总月份数/12）'].sum() / 12, 2)

    basic_salary = int(round(df['基本工资'].sum() / 1000))
    post_salary = int(round(df['岗位工资总'].sum() / 1000))
    scale_salary = int(round(df['薪级工资总'].sum() / 1000))
    perf_pay = int(round(df['绩效'].sum() / 1000))
    nat_sub = int(round(df['国家补贴'].sum() / 1000))
    reform_sub = int(round(df['改革性补贴3'].sum() / 1000))
    tech_reward = int(round(df['酬金中科技成果转化'].sum() / 1000))

    if 'RW01' in wb.sheetnames:
        ws1 = wb['RW01']
        for r in [14, 22]:
            ws1.cell(row=r, column=4, value=1)
            ws1.cell(row=r, column=5, value=num_end)
            ws1.cell(row=r, column=6, value=num_avg).number_format = '0.00'
            ws1.cell(row=r, column=9, value=basic_salary)
            ws1.cell(row=r, column=10, value=perf_pay)
            ws1.cell(row=r, column=11, value=nat_sub)
            ws1.cell(row=r, column=12, value=reform_sub)
            ws1.cell(row=r, column=16, value=tech_reward)

    if 'RW02' in wb.sheetnames:
        ws2 = wb['RW02']
        for r in [24, 25]:
            ws2.cell(row=r, column=3, value=1)
            ws2.cell(row=r, column=4, value=num_end)
            ws2.cell(row=r, column=7, value=num_end)
            ws2.cell(row=r, column=10, value=num_avg).number_format = '0.00'
            ws2.cell(row=r, column=13, value=post_salary)
            ws2.cell(row=r, column=14, value=scale_salary)
            ws2.cell(row=r, column=15, value=perf_pay)
            ws2.cell(row=r, column=16, value=nat_sub)
            ws2.cell(row=r, column=17, value=reform_sub)
            ws2.cell(row=r, column=21, value=tech_reward)


def process_rw03_04(df, wb):
    mgmt_mapping_post = {
        '管理一级': 16, '管理二级': 17, '管理三级': 18, '管理四级': 19,
        '管理五级': 20, '管理六级': 21, '管理七级': 22, '管理八级': 23,
        '管理九级': 24, '管理十级': 25, '管理见习期': 26
    }
    mgmt_mapping_rank = {
        '管理一级': 28, '管理二级': 29, '管理三级': 30, '管理四级': 31,
        '管理五级': 32, '管理六级': 33, '管理七级': 34, '管理八级': 35,
        '管理九级': 36, '管理十级': 37
    }
    tech_mapping = {
        '专技一级': 15, '专技二级': 16, '专技三级': 17, '专技四级': 18,
        '专技五级': 19, '专技六级': 20, '专技七级': 21, '专技八级': 22,
        '专技九级': 23, '专技十级': 24, '专技十一级': 25, '专技十二级': 26,
        '专技十三级': 27, '专技见习期': 28
    }
    worker_mapping = {
        '工勤一级': 30, '工勤二级': 31, '工勤三级': 32,
        '工勤四级': 33, '工勤五级': 34, '工勤普通工': 35
    }

    mgmt_sheets = ['RW03.31', 'RW04.11']
    tech_worker_sheets = ['RW03.32', 'RW04.12']

    grouped = df.groupby(df.iloc[:, 17])

    for post, group in grouped:
        if pd.isna(post): continue

        num_end = int(group['年末的人'].sum())
        num_avg = round(group['总月份数（正式职工年平均人数=总月份数/12）'].sum() / 12, 2)

        post_salary = int(round(group['岗位工资总'].sum() / 1000))
        scale_salary = int(round(group['薪级工资总'].sum() / 1000))
        perf_pay = int(round(group['绩效'].sum() / 1000))
        nat_sub = int(round(group['国家补贴'].sum() / 1000))
        reform_sub = int(round(group['改革性补贴3'].sum() / 1000))
        tech_reward = int(round(group['酬金中科技成果转化'].sum() / 1000))

        if num_end == 0 and num_avg == 0 and post_salary == 0:
            continue

        def write_row(ws_list, row_idx):
            for sheet_name in ws_list:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    ws.cell(row=row_idx, column=6, value=num_end)
                    ws.cell(row=row_idx, column=7, value=num_avg).number_format = '0.00'
                    ws.cell(row=row_idx, column=10, value=post_salary)
                    ws.cell(row=row_idx, column=11, value=scale_salary)
                    ws.cell(row=row_idx, column=12, value=perf_pay)
                    ws.cell(row=row_idx, column=13, value=nat_sub)
                    ws.cell(row=row_idx, column=14, value=reform_sub)
                    ws.cell(row=row_idx, column=18, value=tech_reward)

        if post in mgmt_mapping_post:
            write_row(mgmt_sheets, mgmt_mapping_post[post])
            if post in mgmt_mapping_rank:
                write_row(mgmt_sheets, mgmt_mapping_rank[post])
        elif post in tech_mapping:
            write_row(tech_worker_sheets, tech_mapping[post])
        elif post in worker_mapping:
            write_row(tech_worker_sheets, worker_mapping[post])


def process_rw07_08(df, wb):
    mgmt_mapping_post = {
        '管理一级': 12, '管理二级': 13, '管理三级': 14, '管理四级': 15,
        '管理五级': 16, '管理六级': 17, '管理七级': 18, '管理八级': 19,
        '管理九级': 20, '管理十级': 21, '管理见习期': 22
    }
    mgmt_mapping_rank = {
        '管理一级': 24, '管理二级': 25, '管理三级': 26, '管理四级': 27,
        '管理五级': 28, '管理六级': 29, '管理七级': 30, '管理八级': 31,
        '管理九级': 32, '管理十级': 33
    }
    tech_mapping = {
        '专技一级': 12, '专技二级': 13, '专技三级': 14, '专技四级': 15,
        '专技五级': 16, '专技六级': 17, '专技七级': 18, '专技八级': 19,
        '专技九级': 20, '专技十级': 21, '专技十一级': 22, '专技十二级': 23,
        '专技十三级': 24, '专技见习期': 25
    }

    counts = defaultdict(int)

    # 【新增限制】：只筛选年末的人为1的数据参与人数统计
    df_filtered = df[pd.to_numeric(df['年末的人'], errors='coerce').fillna(0) == 1]

    for index, row in df_filtered.iterrows():
        pos = row.iloc[17]
        scale = row['薪级']

        if pd.isna(pos) or pd.isna(scale): continue
        scale = int(scale)

        if 1 <= scale <= 16:
            sheet_idx = 0
            col_idx = scale + 3
        elif 17 <= scale <= 33:
            sheet_idx = 1
            col_idx = scale - 14
        elif 34 <= scale <= 50:
            sheet_idx = 2
            col_idx = scale - 31
        elif 51 <= scale <= 65:
            sheet_idx = 3
            col_idx = scale - 48
        else:
            continue

        if pos in mgmt_mapping_post:
            target_sheets = ['RW07', 'RW07.1', 'RW07.2', 'RW07.3']
            sheet_name = target_sheets[sheet_idx]
            counts[(sheet_name, mgmt_mapping_post[pos], col_idx)] += 1
            if pos in mgmt_mapping_rank:
                counts[(sheet_name, mgmt_mapping_rank[pos], col_idx)] += 1

        elif pos in tech_mapping:
            target_sheets = ['RW08', 'RW08.1', 'RW08.2', 'RW08.3']
            sheet_name = target_sheets[sheet_idx]
            counts[(sheet_name, tech_mapping[pos], col_idx)] += 1

    for (sheet, r, c), val in counts.items():
        if val > 0 and sheet in wb.sheetnames:
            wb[sheet].cell(row=r, column=c, value=val)


def process_rw09(df, wb):
    role_mapping = {
        '管理一级': ('RW09', 12, True), '管理二级': ('RW09', 16, True),
        '管理三级': ('RW09', 20, True), '管理四级': ('RW09', 24, True),
        '管理五级': ('RW09', 28, True), '管理六级': ('RW09.1', 11, True),
        '管理七级': ('RW09.1', 15, True), '管理八级': ('RW09.1', 19, True),
        '管理九级': ('RW09.1', 23, True), '管理十级': ('RW09.1', 27, False),
        '管理见习期': ('RW09.1', 28, False), '专技一级': ('RW09.4', 13, True),
        '专技二级': ('RW09.4', 17, True), '专技三级': ('RW09.4', 21, True),
        '专技四级': ('RW09.4', 25, True), '专技五级': ('RW09.4', 29, True),
        '专技六级': ('RW09.4', 33, True), '专技七级': ('RW09.5', 11, True),
        '专技八级': ('RW09.5', 15, True), '专技九级': ('RW09.5', 19, True),
        '专技十级': ('RW09.5', 23, True), '专技十一级': ('RW09.5', 27, True),
        '专技十二级': ('RW09.5', 31, True), '专技十三级': ('RW09.5', 35, False),
        '专技见习期': ('RW09.5', 36, False), '工勤一级': ('RW09.6', 13, True),
        '工勤二级': ('RW09.6', 17, True), '工勤三级': ('RW09.6', 21, True),
        '工勤四级': ('RW09.6', 25, True), '工勤五级': ('RW09.6', 29, True),
        '工勤普通工': ('RW09.6', 34, True)
    }
    tenure_offset = {
        '5年以下': 0, '6-10年': 1, '11-15年': 2, '16年及以上': 3
    }
    work_years_col = {
        '5年以下': 5, '6-10年': 6, '11-15年': 7, '16-20年': 8,
        '21-25年': 9, '26-30年': 10, '31-35年': 11, '36-40年': 12,
        '41年以上': 13
    }

    counts = defaultdict(int)

    # 【新增限制】：只筛选年末的人为1的数据参与人数统计
    df_filtered = df[pd.to_numeric(df['年末的人'], errors='coerce').fillna(0) == 1]

    for index, row in df_filtered.iterrows():
        pos = row.iloc[17]
        work_year_cat = row.iloc[16]
        tenure_cat = row.iloc[19]

        if pd.isna(pos) or pd.isna(work_year_cat) or pd.isna(tenure_cat): continue
        if pos not in role_mapping: continue

        sheet_name, base_row, splits = role_mapping[pos]

        if splits:
            if tenure_cat not in tenure_offset: continue
            r = base_row + tenure_offset[tenure_cat]
        else:
            r = base_row

        if work_year_cat not in work_years_col: continue

        c = work_years_col[work_year_cat]
        counts[(sheet_name, r, c)] += 1

    for (sheet, r, c), val in counts.items():
        if val > 0 and sheet in wb.sheetnames:
            wb[sheet].cell(row=r, column=c, value=val)


def process_rw13(df, wb):
    worker_mapping = {
        '工勤一级': 12, '工勤二级': 13, '工勤三级': 14,
        '工勤四级': 15, '工勤五级': 16, '工勤普通工': 17
    }
    counts = defaultdict(int)

    # 【新增限制】：只筛选年末的人为1的数据参与人数统计
    df_filtered = df[pd.to_numeric(df['年末的人'], errors='coerce').fillna(0) == 1]

    for index, row in df_filtered.iterrows():
        pos_raw = row.get('聘任岗位')
        scale_raw = row.get('薪级')

        if pd.isna(pos_raw) or pd.isna(scale_raw):
            continue

        pos = str(pos_raw).strip()

        if pos not in worker_mapping:
            continue

        try:
            scale = int(float(scale_raw))
        except (ValueError, TypeError):
            continue

        row_idx = worker_mapping[pos]

        if 1 <= scale <= 20:
            sheet_name = 'RW13'
            col_idx = scale + 3
        elif 21 <= scale <= 39:
            sheet_name = 'RW13.1'
            col_idx = scale - 18
        elif scale >= 40:
            sheet_name = 'RW13.1'
            col_idx = 22
        else:
            continue

        counts[(sheet_name, row_idx, col_idx)] += 1

    for (sheet, r, c), val in counts.items():
        if val > 0 and sheet in wb.sheetnames:
            wb[sheet].cell(row=r, column=c, value=val)


def process_rw16_17(df, wb):
    sheet_16 = next((s for s in wb.sheetnames if 'RW16' in s), None)
    sheet_17 = next((s for s in wb.sheetnames if 'RW17.1' in s or 'RW17' in s), None)

    col_E = df.columns[4]
    mask_16 = df[col_E].astype(str).str.replace(' ', '').str.contains('事业年薪制', na=False)
    df_rw16 = df[mask_16]

    if not df_rw16.empty and sheet_16:
        ws16 = wb[sheet_16]

        col_end = next((c for c in df.columns if '年末' in str(c)), df.columns[29])
        col_month = next((c for c in df.columns if '总月份' in str(c) and '事业' not in str(c)), df.columns[7])
        col_income = next((c for c in df.columns if '总收入3' in str(c) and '事业' not in str(c)), df.columns[30])

        num_end_16 = int(pd.to_numeric(df_rw16[col_end], errors='coerce').fillna(0).sum())
        num_avg_16 = round(pd.to_numeric(df_rw16[col_month], errors='coerce').fillna(0).sum() / 12, 2)

        salary_total_16 = int(round(pd.to_numeric(df_rw16[col_income], errors='coerce').fillna(0).sum() / 1000))

        for r in [17]:
            ws16.cell(row=r, column=4, value=num_end_16)
            ws16.cell(row=r, column=5, value=num_avg_16).number_format = '0.00'
            ws16.cell(row=r, column=6, value=salary_total_16)

    col_K = df.columns[10]
    df_rw17 = df[pd.to_numeric(df[col_K], errors='coerce').fillna(0) > 0]

    if not df_rw17.empty and sheet_17:
        ws17 = wb[sheet_17]
        col_month = next((c for c in df.columns if '总月份' in str(c) and '事业' not in str(c)), df.columns[7])

        num_avg_17 = round(pd.to_numeric(df_rw17[col_month], errors='coerce').fillna(0).sum() / 12, 2)

        tech_income_17 = int(round(pd.to_numeric(df_rw17[col_K], errors='coerce').fillna(0).sum() / 1000))

        for r in [13, 21]:
            ws17.cell(row=r, column=5, value=1)
            ws17.cell(row=r, column=6, value=num_avg_17).number_format = '0.00'
            ws17.cell(row=r, column=7, value=tech_income_17)


# --- 【全新改版：补齐 RW20 第6行和第11行所有明细】 ---
def process_rw20_series(df, wb):
    col_post = next((c for c in df.columns if '聘任岗位' in str(c)), df.columns[17])
    col_type = next((c for c in df.columns if '分类' in str(c)), df.columns[4])
    col_month = next((c for c in df.columns if '总月份数' in str(c)), df.columns[7])
    col_income = next((c for c in df.columns if '总收入3' in str(c)), df.columns[30])

    col_post_sal = '岗位工资总'
    col_scale_sal = '薪级工资总'
    col_perf = '绩效'
    col_reform = '改革性补贴3'
    col_nat = '国家补贴'
    col_tech = '酬金中科技成果转化'

    # ================= RW20 逻辑 =================
    sheet_20 = next((s for s in wb.sheetnames if 'RW20' in s and '.1' not in s), None)
    if sheet_20:
        ws20 = wb[sheet_20]

        is_career = df[col_type].astype(str).str.contains('事业', na=False)
        df_career = df[is_career]
        df_other = df[~is_career]

        # 辅助计算函数 (获取数据元/千元总额)
        def get_sum_yuan(data, col):
            return pd.to_numeric(data[col], errors='coerce').fillna(0).sum()

        def get_sum_k(data, col):
            return round(get_sum_yuan(data, col) / 1000, 2)

        # 1. 基础总人数计算
        avg_total = round(df[col_month].sum() / 12, 2)
        avg_career = round(df_career[col_month].sum() / 12, 2)
        avg_other = round(df_other[col_month].sum() / 12, 2)

        # 填入平均人数相关列 (第6行，列 1/2/4)
        ws20.cell(row=6, column=1, value=avg_total).number_format = '0.00'
        ws20.cell(row=6, column=2, value=avg_career).number_format = '0.00'
        ws20.cell(row=6, column=4, value=avg_other).number_format = '0.00'

        # 2. 补齐【第6行：各项年平均工资总额（元/人）】
        if avg_total > 0:
            total_income_y = get_sum_yuan(df, col_income)
            post_sal_y = get_sum_yuan(df, col_post_sal)
            scale_sal_y = get_sum_yuan(df, col_scale_sal)
            perf_y = get_sum_yuan(df, col_perf)
            reform_y = get_sum_yuan(df, col_reform)
            nat_y = get_sum_yuan(df, col_nat)
            tech_y = get_sum_yuan(df, col_tech)

            ws20.cell(row=6, column=6, value=round(total_income_y / avg_total, 2)).number_format = '0.00'
            ws20.cell(row=6, column=7, value=round((post_sal_y + scale_sal_y) / avg_total, 2)).number_format = '0.00'
            ws20.cell(row=6, column=8, value=round(post_sal_y / avg_total, 2)).number_format = '0.00'
            ws20.cell(row=6, column=9, value=round(scale_sal_y / avg_total, 2)).number_format = '0.00'
            ws20.cell(row=6, column=10, value=round(perf_y / avg_total, 2)).number_format = '0.00'
            ws20.cell(row=6, column=13, value=round(reform_y / avg_total, 2)).number_format = '0.00'
            ws20.cell(row=6, column=14, value=round(nat_y / avg_total, 2)).number_format = '0.00'
            ws20.cell(row=6, column=15, value=round(tech_y / avg_total, 2)).number_format = '0.00'

        # 3. 补齐【第11行：年工资总额各项明细（千元）】
        ws20.cell(row=11, column=1, value=get_sum_k(df, col_income)).number_format = '0.00'
        ws20.cell(row=11, column=4,
                  value=get_sum_k(df, col_post_sal) + get_sum_k(df, col_scale_sal)).number_format = '0.00'
        ws20.cell(row=11, column=7, value=get_sum_k(df, col_post_sal)).number_format = '0.00'
        ws20.cell(row=11, column=8, value=get_sum_k(df, col_scale_sal)).number_format = '0.00'
        ws20.cell(row=11, column=9, value=get_sum_k(df, col_perf)).number_format = '0.00'
        ws20.cell(row=11, column=12, value=get_sum_k(df, col_nat)).number_format = '0.00'
        ws20.cell(row=11, column=13, value=get_sum_k(df, col_reform)).number_format = '0.00'
        ws20.cell(row=11, column=14, value=get_sum_k(df, col_tech)).number_format = '0.00'

    # ================= RW20.1 逻辑 =================
    sheet_201 = next((s for s in wb.sheetnames if 'RW20.1' in s), None)
    if sheet_201:
        ws201 = wb[sheet_201]

        post_row_map = {
            '管理二级': 2, '管理三级': 3, '管理四级': 4, '管理五级': 5, '管理六级': 6,
            '管理七级': 7, '管理八级': 8, '管理九级': 9, '管理十级': 10, '管理见习期': 11,
            '专技一级': 13, '专技二级': 14, '专技三级': 15, '专技四级': 16, '专技五级': 17,
            '专技六级': 18, '专技七级': 19, '专技八级': 20, '专技九级': 21, '专技十级': 22,
            '专技十一级': 23, '专技十二级': 24, '专技十三级': 25, '专技见习期': 26,
            '工勤一级': 28, '工勤二级': 29, '工勤三级': 30, '工勤四级': 31, '工勤五级': 32, '工勤普通工': 33
        }

        def write_rw201_row(row_idx, subset_df):
            if subset_df.empty: return

            total_months = pd.to_numeric(subset_df[col_month], errors='coerce').fillna(0).sum()
            avg_num = round(total_months / 12, 2)
            total_inc = round(pd.to_numeric(subset_df[col_income], errors='coerce').fillna(0).sum() / 1000, 2)
            per_capita = round((total_inc * 1000 / avg_num), 2) if avg_num > 0 else 0

            c_subset = subset_df[subset_df[col_type].astype(str).str.contains('事业', na=False)]
            c_months = pd.to_numeric(c_subset[col_month], errors='coerce').fillna(0).sum()
            c_avg_num = round(c_months / 12, 2)
            c_total_inc = round(pd.to_numeric(c_subset[col_income], errors='coerce').fillna(0).sum() / 1000, 2)
            c_per_capita = round((c_total_inc * 1000 / c_avg_num), 2) if c_avg_num > 0 else 0

            data_cells = [
                (2, total_months), (3, avg_num), (4, total_inc), (5, per_capita),
                (6, c_months), (7, c_avg_num), (8, c_total_inc), (9, c_per_capita)
            ]
            for col_idx, val in data_cells:
                ws201.cell(row=row_idx, column=col_idx, value=val).number_format = '0.00'

        grouped = df.groupby(col_post)
        for post_name, g in grouped:
            if pd.notna(post_name) and post_name in post_row_map:
                write_rw201_row(post_row_map[post_name], g)

        df_mgmt = df[df[col_post].astype(str).str.contains('管理', na=False)]
        write_rw201_row(12, df_mgmt)

        df_tech = df[df[col_post].astype(str).str.contains('专技', na=False)]
        write_rw201_row(27, df_tech)

        df_worker = df[df[col_post].astype(str).str.contains('工勤', na=False)]
        write_rw201_row(34, df_worker)

        write_rw201_row(35, df)


# ================= GUI 主程序 =================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("RW 表格填报汇总工具")
        self.root.geometry("500x250")
        self.root.resizable(False, False)

        self.data_file_path = tk.StringVar()
        self.template_file_path = tk.StringVar()

        label_font = ('微软雅黑', 10)
        btn_font = ('微软雅黑', 9)

        tk.Label(root, text="第一步：选择导入数据表 (Excel)", font=label_font).pack(anchor="w", padx=20, pady=(15, 5))
        frame1 = tk.Frame(root)
        frame1.pack(fill="x", padx=20)
        tk.Entry(frame1, textvariable=self.data_file_path, state="readonly", width=45).pack(side="left", ipady=3)
        tk.Button(frame1, text="浏览...", font=btn_font, command=self.select_data).pack(side="right", padx=5)

        tk.Label(root, text="第二步：选择填报模板 (Excel)", font=label_font).pack(anchor="w", padx=20, pady=(15, 5))
        frame2 = tk.Frame(root)
        frame2.pack(fill="x", padx=20)
        tk.Entry(frame2, textvariable=self.template_file_path, state="readonly", width=45).pack(side="left", ipady=3)
        tk.Button(frame2, text="浏览...", font=btn_font, command=self.select_template).pack(side="right", padx=5)

        self.run_btn = tk.Button(root, text="开始处理并写入", font=('微软雅黑', 11, 'bold'), bg="#4CAF50", fg="white",
                                 command=self.run_processing, height=2)
        self.run_btn.pack(fill="x", padx=20, pady=25)

    def select_data(self):
        path = filedialog.askopenfilename(title="选择明细数据表", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if path: self.data_file_path.set(path)

    def select_template(self):
        path = filedialog.askopenfilename(title="选择样表模板", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if path: self.template_file_path.set(path)

    def run_processing(self):
        data_path = self.data_file_path.get()
        template_path = self.template_file_path.get()

        if not data_path or not template_path:
            messagebox.showwarning("提示", "请先选择数据源文件和模板文件！")
            return

        self.run_btn.config(text="处理中，请稍候...", state="disabled", bg="#A5D6A7")
        self.root.update()

        try:
            df = pd.read_excel(data_path)
            wb = load_workbook(template_path)

            process_rw01_02(df, wb)
            process_rw03_04(df, wb)
            process_rw07_08(df, wb)
            process_rw09(df, wb)
            process_rw13(df, wb)
            process_rw16_17(df, wb)

            # 【完美覆盖】执行最新版本的 RW20 & RW20.1 处理逻辑
            process_rw20_series(df, wb)

            wb.save(template_path)
            messagebox.showinfo("成功",
                                f"✅ 所有报表处理完毕并保存至：\n{template_path}\n\n所有金额换算已全面替换为标准的四舍五入！")

        except PermissionError:
            messagebox.showerror("错误", "保存失败：模板文件正在被其它程序（如Excel）占用，请关闭该文件后重试！")
        except Exception as e:
            err_msg = traceback.format_exc()
            messagebox.showerror("运行异常", f"处理过程中发生错误：\n{str(e)}\n\n详情请见控制台。")
            print(err_msg)
        finally:
            self.run_btn.config(text="开始处理并写入", state="normal", bg="#4CAF50")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
